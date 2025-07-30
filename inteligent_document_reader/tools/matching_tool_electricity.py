"""
Electricity Matching Tool for Excel validation - reads input Excel and writes output Excel with electricity matches
"""

import json
import pandas as pd
from pathlib import Path
from typing import Any
from langchain.tools import BaseTool
from pydantic import Field

class MatchingToolElectricity(BaseTool):
    """Tool for matching Excel data with electricity documents and outputting to Excel format"""
   
    name: str = "excel_electricity_matching_tool"
    description: str = "Reads input Excel file (rows 6+, columns B-J), matches locations with electricity database, writes results to output_electricity.xlsm (columns K-Q). Input should be JSON with 'input_excel_path'."
    search_client: Any = Field(default=None, exclude=True)
   
    def __init__(self, search_client, **kwargs):
        super().__init__(**kwargs)
        object.__setattr__(self, 'search_client', search_client)
   
    def _search_electricity_by_location(self, location: str):
        """Search electricity database for location matches"""
        try:
            print(f"   🔍 Searching electricity database for: {location}")
           
            # Search for electricity documents with matching location
            search_results = self.search_client.search(
                search_text=f"{location}",
                filter="doc_type eq 'Electricity'",  # Only electricity documents
                select=["id", "filename", "location", "client_name", "start_date",
                       "end_date", "consumption", "consumption_unit", "cost", "cost_unit", "processed_at"],
                top=50
            )
           
            # Also try filter-based search for exact client matches
            try:
                filtered_results = self.search_client.search(
                    search_text="*",
                    filter=f"doc_type eq 'Electricity' and client_name eq '{location}'",
                    select=["id", "filename", "location", "client_name", "start_date",
                           "end_date", "consumption", "consumption_unit", "cost", "cost_unit", "processed_at"],
                    top=50
                )
                all_results = list(search_results) + list(filtered_results)
            except:
                all_results = list(search_results)
           
            # Process results and find best match
            matches = []
            for result in all_results:
                matches.append({
                    "filename": result.get("filename"),
                    "location": result.get("location"),
                    "client_name": result.get("client_name"),
                    "start_date": result.get("start_date"),
                    "end_date": result.get("end_date"),
                    "consumption": result.get("consumption"),
                    "consumption_unit": result.get("consumption_unit"),
                    "cost": result.get("cost"),
                    "cost_unit": result.get("cost_unit"),
                    "score": result.get("@search.score", 0)
                })
           
            # Group by filename to avoid counting pages as separate documents
            documents = {}
            for match in matches:
                filename = match.get("filename", "Unknown")
                if filename not in documents:
                    documents[filename] = match  # Take first page data for the document
           
            unique_matches = list(documents.values())
           
            # Return best match (highest score)
            if unique_matches:
                best_match = max(unique_matches, key=lambda x: x.get("score", 0))
                print(f"      ✅ Found {len(unique_matches)} matches, best: {best_match['filename']}")
                return best_match
            else:
                print(f"      ❌ No matches found")
                return None
           
        except Exception as e:
            print(f"   ❌ Error searching for location: {e}")
            return None
   
    def _run(self, input_data: str) -> str:
        """Process Excel file and create output with electricity matches"""
        try:
            data = json.loads(input_data)
            input_excel_path = data.get('input_excel_path', 'Electricity_example.xlsm')
            output_excel_path = data.get('output_excel_path', 'output_electricity.xlsm')
           
            print(f"📊 ELECTRICITY MATCHING TOOL")
            print(f"   Input: {input_excel_path}")
            print(f"   Output: {output_excel_path}")
           
            # Read the input Excel file
            try:
                # Read with no header to access raw structure
                df_raw = pd.read_excel(input_excel_path, header=None, engine='openpyxl')
                print(f"   📈 Loaded Excel file: {df_raw.shape}")
            except Exception as e:
                return f"❌ Error loading Excel file '{input_excel_path}': {e}"
           
            # Copy the entire structure to output
            df_output = df_raw.copy()
           
            # Find data rows (starting from row 6, which is index 5)
            data_start_row = 5  # Row 6 (0-indexed as 5)
           
            if len(df_output) <= data_start_row:
                return f"❌ Excel file has no data rows. Expected data starting from row 6."
           
            processed_rows = 0
            matched_rows = 0
           
            # Process each data row
            for row_idx in range(data_start_row, len(df_output)):
                # Read input data from columns B-J (indices 1-9)
                location = df_output.iloc[row_idx, 2]  # Column C (index 2)
               
                # Skip rows with no location data
                if pd.isna(location) or str(location).strip() == '':
                    continue
               
                processed_rows += 1
                print(f"
📍 Row {row_idx + 1}: Processing location '{location}'")
               
                # Search for electricity matches
                electricity_match = self._search_electricity_by_location(str(location))
               
                if electricity_match:
                    matched_rows += 1
                    print(electricity_match)
                    # Write electricity data to columns K-Q (indices 10-16)
                    df_output.iloc[row_idx, 10] = "Electricity Bill"  # Column K
                    df_output.iloc[row_idx, 11] = electricity_match.get('location', '')  # Column L
                    df_output.iloc[row_idx, 12] = electricity_match.get('start_date', '')  # Column M
                    df_output.iloc[row_idx, 13] = electricity_match.get('end_date', '')  # Column N
                    df_output.iloc[row_idx, 14] = electricity_match.get('consumption', '')  # Column O
                    df_output.iloc[row_idx, 15] = electricity_match.get('consumption_unit', '')  # Column P
                    df_output.iloc[row_idx, 16] = electricity_match.get('cost', '')  # Column Q
                    df_output.iloc[row_idx, 17] = electricity_match.get('cost_unit', '') # Column R
 
                    print(f"   ✅ Matched with electricity: {electricity_match.get('filename', 'Unknown')}")
                    print(f"      Location: {electricity_match.get('location', 'N/A')}")
                    print(f"      Period: {electricity_match.get('start_date', 'N/A')} to {electricity_match.get('end_date', 'N/A')}")
                    print(f"      Consumption: {electricity_match.get('consumption', 'N/A')} {electricity_match.get('consumption_unit', '')}")
                    print(f"      Cost: {electricity_match.get('cost', 'N/A')} {electricity_match.get('cost_unit', '')}")
                else:
                    # Clear columns K-Q if no match
                    for col_idx in range(10, 18):
                        df_output.iloc[row_idx, col_idx] = ''
                    print(f"   ❌ No electricity match found")
           
            # Save the output Excel file
            try:
                # Use xlsxwriter for better Excel compatibility, or keep_vba for .xlsm files
                if output_excel_path.endswith('.xlsm'):
                    # For .xlsm files, we need to preserve the format
                    from openpyxl import load_workbook
                   
                    # Load the original workbook to preserve formatting and macros
                    wb = load_workbook(input_excel_path, keep_vba=True)
                    ws = wb.active
                   
                    # Update the worksheet with our data
                    for row_idx in range(len(df_output)):
                        for col_idx in range(len(df_output.columns)):
                            # Excel is 1-indexed, pandas is 0-indexed
                            cell_value = df_output.iloc[row_idx, col_idx]
                            if pd.notna(cell_value):
                                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                   
                    # Save with VBA preserved
                    wb.save(output_excel_path)
                else:
                    # For regular .xlsx files
                    df_output.to_excel(output_excel_path, header=False, index=False, engine='openpyxl')
               
                print(f"
💾 Saved output to: {output_excel_path}")
            except Exception as e:
                return f"❌ Error saving output file: {e}"
           
            # Summary
            result = f"📊 ELECTRICITY MATCHING COMPLETED
"
            result += f"{'='*40}
"
            result += f"📁 Input file: {input_excel_path}
"
            result += f"📁 Output file: {output_excel_path}
"
            result += f"📈 Summary:
"
            result += f"   • Rows processed: {processed_rows}
"
            result += f"   • Rows with electricity matches: {matched_rows}
"
            result += f"   • Match rate: {(matched_rows/processed_rows*100):.1f}% 
" if processed_rows > 0 else "   • Match rate: 0%
"
            result += f"   • Output columns: K-R populated with electricity data
"
            result += f"
✅ Results written to {output_excel_path}"
           
            return result
           
        except json.JSONDecodeError:
            return "❌ Error: Invalid JSON input format. Use: {'input_excel_path': 'Electricity_example.xlsm'}"
        except Exception as e:
            return f"❌ Error processing Excel matching: {str(e)}"
   
    async def _arun(self, input_data: str) -> str:
        """Use the tool asynchronously."""
        raise NotImplementedError()
